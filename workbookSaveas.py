import os
from shutil import copyfile
import datetime
import time
from tkinter import *
from tkinter import filedialog, messagebox

# import threading

import win32gui, win32api, win32con

from openpyxl import Workbook,load_workbook


def openFolderOrFile(txt_FolderOrFile):
    path_FolderOrFile=txt_FolderOrFile.get('0.0',END).strip()
    if os.path.exists(path_FolderOrFile):
        # 打开文件夹或文件
        os.startfile(path_FolderOrFile)


# --函数：选择文件或文件夹显示到文本框--
def selectFileOrFolder(myTextbox,txt_TemplateFileName=None,isFile=True):
    url=''
    if isFile:
        url = filedialog.askopenfilename(title='选择Excel文件',filetypes=(('Excel工作簿','*.xlsx'),))
        
        if url!='':
            index1 = url.rindex('/')+1
            filename=url[index1:]
            txt_TemplateFileName['text']='运行中的-'+filename
    else:
        url = filedialog.askdirectory()

    if url!='':
        myTextbox['state']='normal'
        myTextbox.delete('0.0',END)
        myTextbox.insert('0.0',url)
        myTextbox['state']='disabled'

# --函数：复制文件，isOpenFile是否打开文件--
def funCopyFile(path_sourceFile,path_targetFile,isOpenFile=False):
    # 复制文件到保存文件夹
    try:
        copyfile(path_sourceFile, path_targetFile)
    except IOError as e:
        messagebox.showerror('不能复制文件：','%s' %e)
        raise IOError('不能复制文件：%s' %e) # 抛出异常

    except Exception as e:
        messagebox.showerror('复制文件的其他异常:', str(e))
        raise IOError('复制文件的其他异常:'+ str(e))
    else:
        # 么有异常就打开文件
        if isOpenFile:
            # 打开复制后的文件
            os.startfile(path_targetFile)

# --函数：根据无扩展名的文件名，获取窗口句柄、标题--
def funGetWindowHandler(filename_noExtension):
    map_hwnd_title=dict()

    # 获取窗口句柄
    def get_all_windowHandler(hwnd,mouse):
        if win32gui.IsWindow(hwnd) and win32gui.IsWindowEnabled(hwnd) and win32gui.IsWindowVisible(hwnd):
            map_hwnd_title[hwnd]=win32gui.GetWindowText(hwnd)

    # 列举出所有窗口的句柄
    win32gui.EnumWindows(get_all_windowHandler,0)

    list_handlerAndTitle=[]
    for winHandler,winTitle in map_hwnd_title.items():
    
        if winTitle.strip()[0:len(filename_noExtension)]==filename_noExtension:
            list_handlerAndTitle=[winHandler,winTitle]
            break
    return list_handlerAndTitle


def setRunInfo(step_run,setp_total,str_info,myTextbox,isError=False,isStart=True):
    str_step=str(step_run) +'/'+str(setp_total)
    if isStart:
        str_step='第'+str_step+'步：'
    else:
        str_step='已完成'+str_step+'步：'

    str_result=str_step+str_info

    # myTextbox.config(state='normal') 等效与 myTextbox['state']='normal'
    myTextbox['state']='normal'
    myTextbox.delete('0.0',END)
    myTextbox.insert('0.0',str_result)
    myTextbox['state']='disabled'

    myTextbox['height']=getStrWrapLines(str_result,myTextbox['width'])

    if isError:
        myTextbox['bg']='pink'
    else:
        if myTextbox['bg']!='lightgreen':
            myTextbox['bg']='lightgreen'

    myTextbox.update()

def getStrWrapLines(str,width_line):
    list_line=str.split('\n')
    line_count=0
    for line in list_line:
        chrLen=strLenToChrLen(line)
        if chrLen>width_line:
            line_count +=chrLen//width_line
            if chrLen % width_line != 0:
                line_count+=1
        else:
            line_count+=1
    return line_count


global isRunning
isRunning=False
# --函数：工作簿另存为--
def workbookSaveas(dict_widget):
    try:
        global isRunning
        if isRunning:return

        # 获取参数字典中的值
        fm_map=dict_widget['fm_map']
        txt_runInfo=dict_widget['txt_runInfo']
        txt_srcFile=dict_widget['txt_srcFile']
        txt_targetFolder=dict_widget['txt_targetFolder']
        txt_sht_name=dict_widget['txt_sht_name']
        txt_filename_cellAddr=dict_widget['txt_filename_cellAddr']
        txt_orderNumber_cellAddr=dict_widget['txt_orderNumber_cellAddr']
        # txt_orderTime_cellAddr=dict_widget['txt_orderTime_cellAddr']

        path_srcFile = txt_srcFile.get('0.0',END).strip()
        path_targetFolder=txt_targetFolder.get('0.0',END).strip()

        if os.path.isfile(path_srcFile)==False or os.path.isdir(path_targetFolder)==False:
            # 源文件和目标文件夹有一个不存在，就退出
            isRunning=False
            return

        # os.path判断'/'的相对路径为True
        path_targetFolder+='/'

        # 文件、文件夹都存在时，运行
        isRunning=True

        fm_map.grid_remove()    # 隐藏控件
        fm_map.master.update()

        index_filenameStart = path_srcFile.rindex('/')+1
        index_dot =path_srcFile.rindex('.')

        extension_filename=path_srcFile[index_dot:]

        prefix_filename='运行中的-'

        filename_noExtension = prefix_filename + path_srcFile[index_filenameStart:index_dot]

        path_runTemplateFile = path_targetFolder + prefix_filename + path_srcFile[index_filenameStart:]

        setp_total=7
        step_run=1

        str_info='获取运行中模板工作簿的窗口：'+  prefix_filename + path_srcFile[index_filenameStart:]
        setRunInfo(step_run,setp_total,str_info,txt_runInfo)

        list_handlerAndTitle = funGetWindowHandler(filename_noExtension)

        if list_handlerAndTitle:
            # 列表不为空，存在窗口，取出 句柄、窗口标题
            winHandler_runTemplate,winTitle_runTemplate=list_handlerAndTitle
        else:
            str1='请手动编辑“运行中模板”工作簿后，再次点击“另存为”按钮运行！'
            str_info=str1+'\n\n首次运行或之前未打开“运行中模板”时出现此提示。'
            setRunInfo(step_run,setp_total,str_info,txt_runInfo)

            # 没窗口就复制打开模板文件，退出程序
            funCopyFile(path_srcFile,path_runTemplateFile,True)

            setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)
            messagebox.showinfo(str1,str1)

            fm_map.grid(row=2,column=2,sticky='NW') # 显示框架
            fm_map.master.update()

            isRunning=False
            return

        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        dataShtName=txt_sht_name.get().strip()
        filename_cellAddr=txt_filename_cellAddr.get().strip()
        orderNumber_cellAddr=txt_orderNumber_cellAddr.get().strip()
        # orderTime_cellAddr=txt_orderTime_cellAddr.get().strip()


        # -------
        t1_Modify=os.path.getmtime(path_runTemplateFile)    # 保存操作前的修改时间

        # win32gui.SetWindowPos(winHandler_runTemplate, win32con.HWND_TOPMOST, 0, 0, 0, 0,win32con.SWP_NOMOVE | win32con.SWP_NOACTIVATE | win32con.SWP_NOOWNERZORDER | win32con.SWP_SHOWWINDOW | win32con.SWP_NOSIZE)    # 窗口置顶

        if win32gui.IsIconic(winHandler_runTemplate):
            win32gui.ShowWindow(winHandler_runTemplate,win32con.SW_RESTORE) # 取消最小化

        step_run+=1
        str_info='按键ctrl+s保存运行中模板工作簿。'
        setRunInfo(step_run,setp_total,str_info,txt_runInfo)

        win32gui.SetForegroundWindow(winHandler_runTemplate)    # 窗口激活

        # 按键保存  ctrl+S
        win32api.keybd_event(17, 0, 0, 0)   # 按下CTRL
        time.sleep(0.5)
        win32api.PostMessage(winHandler_runTemplate,win32con.WM_KEYDOWN,83)     # S
        win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)    # 放开CTRL

        t_start=datetime.datetime.now().strftime('%H:%M:%S')   # 当前时间

        t_start=time.time() # 当前时间
        while True:
            # 如果文件修改时间改变，退出循环
            if os.path.getmtime(path_runTemplateFile) > t1_Modify: break

            # 循环超过30秒，退出程序
            if (datetime.datetime.fromtimestamp(time.time())-datetime.datetime.fromtimestamp(t_start)).seconds > 30:
                str1='警告：保存文件，运行超时！'
                str_info=str1+'程序已停止。请直接点击“另存为”按钮，重试。\n\n之前信息：'+str_info

                messagebox.showwarning(str1,str_info)
                setRunInfo(step_run,setp_total,str_info,txt_runInfo,isError=True)

                fm_map.grid(row=2,column=2,sticky='NW') # 显示框架
                fm_map.master.update()

                isRunning=False
                return

            fm_map.master.update()
            time.sleep(0.5)

        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        step_run+=1
        str_info='读取运行中模板中的信息：成品名称(要保存的文件名)、工单号、制单时间。'
        setRunInfo(step_run,setp_total,str_info,txt_runInfo)

        # 读取工作簿中的信息
        dict_wbInfo=readWorkbookInfo(path_runTemplateFile,dataShtName,filename_cellAddr,orderNumber_cellAddr)

        # 获取readWorkbookInfo函数返回值中的数据
        new_filenName = dict_wbInfo['new_filenName']

        path_newFile = path_targetFolder + new_filenName+extension_filename

        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        step_run+=1
        str_info='运行模板工作簿另存为：%s。' %path_newFile
        setRunInfo(step_run,setp_total,str_info,txt_runInfo)

        # 复制文件来重命名
        funCopyFile(path_runTemplateFile, path_newFile)

        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        formula_linkNewFile='= HYPERLINK("' + path_newFile + '","' + new_filenName + '")'

        list_data=[dict_wbInfo['orderNumber'],formula_linkNewFile,dict_wbInfo['orderTime']]

        step_run+=1
        str_info='记录操作信息到工作簿，在程序同目录下。'
        setRunInfo(step_run,setp_total,str_info,txt_runInfo)

        # 记录操作信息到工作簿
        recordInfo(list_data)

        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        # 关闭窗口
        # win32gui.PostMessage(winHandler_runTemplate,win32con.WM_CLOSE,3,0) # 此方法不好，会关闭所有excel窗口 # PostMessage异步，SendMessage同步

        step_run+=1
        str_info='按键ctrl+f4关闭运行中模板工作簿。'
        setRunInfo(step_run,setp_total,str_info,txt_runInfo)
        
        win32gui.SetForegroundWindow(winHandler_runTemplate)    # 窗口激活

        # 按键关闭窗口  ctrl+f4 或者 ctrl+w
        win32api.keybd_event(17, 0, 0, 0)   # 按下CTRL
        time.sleep(0.5)
        win32api.PostMessage(winHandler_runTemplate,win32con.WM_KEYDOWN,win32con.VK_F4) #f4
        win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)    # 放开CTRL

        t_start=time.time()
        while True:
            # 没窗口就退出循环
            if win32gui.FindWindow(None,winTitle_runTemplate) == 0: break

            # 循环超过30秒，退出程序
            if (datetime.datetime.fromtimestamp(time.time())-datetime.datetime.fromtimestamp(t_start)).seconds > 30:
                str1='警告：关闭运行中模板，运行超时！'
                str_info=str1+'程序已停止。请直接点击“另存为”按钮，重试。\n\n之前信息：'+str_info

                messagebox.showwarning(str1,str_info)
                setRunInfo(step_run,setp_total,str_info,txt_runInfo,isError=True)

                fm_map.grid(row=2,column=2,sticky='NW') # 显示框架
                fm_map.master.update()

                isRunning=False
                return

            fm_map.master.update()  # 更新界面，防止无响应
            time.sleep(0.5)

        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        step_run+=1
        str_info='自动复制并打开运行中模板，来达到清除表中数据的目的。\n\n运行中模板窗口出现后，请编辑它后，点击“另存为”按钮运行。'
        setRunInfo(step_run,setp_total,str_info,txt_runInfo)

        # 没窗口就复制打开模板文件
        funCopyFile(path_srcFile,path_runTemplateFile,True)

        t_start=time.time()
        while True:
            # 没窗口就退出循环
            if win32gui.FindWindow(None,winTitle_runTemplate) == 0: break

            # 循环超过30秒，退出程序
            if (datetime.datetime.fromtimestamp(time.time())-datetime.datetime.fromtimestamp(t_start)).seconds > 30:
                str1='警告：重新复制并打开运行中模板，运行超时！'
                str_info=str1+'程序已停止。请直接点击“另存为”按钮，重试。\n\n之前信息：'+str_info

                messagebox.showwarning(str1,str_info)
                setRunInfo(step_run,setp_total,str_info,txt_runInfo,isError=True)

                fm_map.grid(row=2,column=2,sticky='NW') # 显示框架
                fm_map.master.update()

                isRunning=False
                return

            fm_map.master.update()  # 更新界面，防止无响应
            time.sleep(0.5)

        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        str_info='运行成功！请手动编辑“运行中模板”后，再次点击“另存为”按钮运行。'
        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isStart=False)

        messagebox.showinfo('运行成功！',str_info)


        txt_runInfo['state']='normal'
        txt_runInfo.delete('0.0',END)
        t_now=datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        str_info='%s %s\n\n%s' %(t_now,str_info,str_tips)
        txt_runInfo.insert('0.0',str_info)
        txt_runInfo['height']=getStrWrapLines(str_info,txt_runInfo['width'])
        txt_runInfo['state']='disabled'
        txt_runInfo['bg']=myColor_win_bg

        fm_map.grid(row=2,column=2,sticky='NW') # 显示框架
        fm_map.master.update()

        isRunning=False


        # -------
        # # 因为可能阻塞，所以创建线程
        # #设置线程方法传入的参数值
        # dict_thread_args = {'strVar_info':strVar_info}

        # myTthreadEvent=threading.Event()
        # myTthreadEvent.clear()

        # #创建线程
        # myTthread = threading.Thread(target = funThreadAction,kwargs =dict_thread_args)
        # myTthread.start()
        # myTthreadEvent.wait()
    except Exception as e:
        str_info='运行错误（%s行）程序已停止：%s。\n\n之前信息：%s' %(e.__traceback__.tb_lineno, e, str_info)
        setRunInfo(step_run,setp_total,str_info,txt_runInfo,isError=True)

        with open(file='log',mode='a',encoding='utf-8') as f:
            t_now=datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
            f.write('%s,%s\n' %(t_now,str_info))

        fm_map.grid(row=2,column=2,sticky='NW') # 显示框架
        fm_map.master.update()

        isRunning=False
        raise Exception(str(e))


# # --函数：线程要调用的函数，**传多个参数--
# def funThreadAction(**dict_thread_args):
    # pass


# --函数：读取工作簿中信息--
def readWorkbookInfo(path_runTemplateFile,dataShtName,filename_cellAddr,orderNumber_cellAddr):

    dict_wbInfo={}

    # openpyxl 获取 newfilename
    wkb = load_workbook(path_runTemplateFile)

    if dataShtName in wkb.sheetnames:
        sht = wkb[dataShtName]
    else:
        str1= dataShtName + ' 表单不存在！（请检查配置）\n\n错误信息：'+str(e)
        messagebox.showerror('获取表单错误:',str1)
        raise ValueError(str1)

    try:
        new_filenName=sht[filename_cellAddr].value
        if new_filenName:
            new_filenName=str(new_filenName)
        else:
            str1='单元格['+filename_cellAddr+']不能为空！请输入后，再点击“另存为”运行。'
            messagebox.showerror('单元格不能为空：',str1)
            raise Exception(str1)

        orderNumber=sht[orderNumber_cellAddr].value
        if orderNumber:
            orderNumber=str(orderNumber)
        else:
            orderNumber=''

        orderTime=datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')   # 当前时间 # %H:%M:%S

    except ValueError as e:
        str1='单元格地址错误（请检查配置）：'+str(e)
        messagebox.showerror('单元格地址错误',str1)
        raise ValueError(str1) # 抛出异常
    except Exception as e:
        str1='读取单元格错误（请检查配置）：'+str(e)
        messagebox.showerror('读取单元格错误:',str1)
        raise Exception(str1)
    else:
        # 设置返回值
        dict_wbInfo['new_filenName']=new_filenName
        dict_wbInfo['orderNumber']=orderNumber
        dict_wbInfo['orderTime']=orderTime
    finally:
        wkb.close()

    return dict_wbInfo


def recordInfo(list_data):
    # path_infoFolder= os.getcwd()    # 当前工作文件夹
    path_infoFile = '记录表-Excel另存为-zhy.xlsx'

    shtName='记录表'

    if not os.path.isfile(path_infoFile):
        # 如果文件不存在就新建
        wkb = Workbook()
        sht = wkb.active
        sht.title = shtName

        # 标题
        sht.append(['工单号','品名','开单时间'])

    else:
         # 如果文件存在就打开
        wkb=load_workbook(path_infoFile)

        if shtName in wkb.sheetnames:
            sht = wkb[shtName]
        else:
            sht = wkb.create_sheet(shtName)

        # 获取最大行
        max_row_real =0
        for xRow in sht.iter_rows():
            for xCell in xRow:
                if xCell.value: max_row_real=xCell.row

        max_row=sht.max_row
        if max_row > max_row_real: sht.delete_rows(max_row_real+1,max_row) # 删除空行

    # 赋值，在表末尾添加行，只能一行行添加
    sht.append(list_data)

    try:
        wkb.save(path_infoFile) # 保存工作簿，保存方法自带关闭文件功能
    except Exception as e:
        str1='保存记录工作簿错误：工作簿可能未关闭：“%s”\n\n错误信息：'%path_infoFile +str(e)
        messagebox.showerror('保存记录工作簿错误：' ,str1)
        raise Exception(str1)


def funMouseMove(event,myWindow):
    myWindow.geometry('+{}+{}'.format(event.x_root-mouseX, event.y_root-30))


def funMouseDown(event):
    global mouseX
    mouseX=event.x


# --显示或隐藏框架--
def switchFrame(myFrame_1,myFrame_2):
    global column_1
    global column_2
    # 互换列号
    column_1,column_2=column_2,column_1
    myFrame_1.grid(column=column_1)
    myFrame_2.grid(column=column_2)

    myFrame_1.master.update()

# 读写配置文件
def readWriteMapping(openMode,list_data=[]):
    myFilename='config'
    if openMode=='r':
        if os.path.isfile(myFilename):
            # open内置函数不是os库里的
            f=open (myFilename,openMode,encoding='utf-8')
            for line in f:
                line=line.strip()   # 去掉换行符
                list_data.append(line.split(','))
            f.close()

        else:
            list_data=[
                        ['保存到文件夹：',''],
                        ['数据所在表单名：','单据'],
                        ['另存为文件名单元格地址：','B4'],
                        ['工单号单元格地址：','C2']
                    ]
        return list_data

    elif openMode=='w':
        if list_data==[]:return
        f=open (myFilename,openMode,encoding='utf-8')
        for line in list_data:
            value_line=str.join(',',line)
            f.write(value_line +'\n')
        f.close()
        return

def getConfigData(list_mapping):
    list_data=[]

    for i in range(0,len(list_mapping)):
        myLbl=list_mapping[i][0]
        value_lbl=myLbl.cget('text')
        value_lbl=str(value_lbl)

        mytextbox=list_mapping[i][1]

        if type(mytextbox)==Entry:
            value_text=mytextbox.get()
        else:
            value_text=mytextbox.get('0.0',END)

        value_text=value_text.strip()

        list_data.append([value_lbl,value_text])
    return list_data


def closeWindow(mainWindow):
    if messagebox.askyesno('是否退出？','是否退出程序？\n\n退出后请手动关闭“运行中模板”。'):
        mainWindow.destroy()
        # mainWindow.quit()


def myItemconfigure(myCanvas,list_itemId,dict):
    for itemId in list_itemId:
        myCanvas.itemconfigure(itemId,dict)

    myCanvas.update()

def showOrHideTips(btn_showOrHideTips,fm_runInfo):
    if btn_showOrHideTips['text']=='隐藏 运行提示':
        fm_runInfo.grid_remove()
        btn_showOrHideTips['text']='显示 运行提示'
    else:
        fm_runInfo.grid(row=3,column=2,sticky='NW')
        btn_showOrHideTips['text']='隐藏 运行提示'

def strLenToChrLen(str):
    # len(str.encode())获取按编码的长度;len(str)获取默认编码下字符串看到的长度
    # utf-8 中文占3字节
    # 转换成1个中文等于2英文字符长度
    zhLen=(len(str.encode('utf-8'))-len(str))/2
    enLen=len(str)-zhLen
    lenToChrLen=2*zhLen+enLen
    return lenToChrLen


global str_tips
str_tips='运行提示：\n\n1.手动选择参数、修改配置（可选），然后点击“另存为”按钮。\n\n（原理：自动复制并打开“运行中模板”工作簿，后面再次打开时相当于清空了输入数据，前提是源文件中没有输入数据）。\n\n2.在“运行中模板”中手动输入信息，然后点击“另存为”按钮。\n\n（原理：自动ctrl+s保存“运行中模板”，自动复制模板重命名为成品名工作簿，自动记录操作到程序同文件夹下的工作簿，自动ctrl+f4关闭模板，再次复制并打开模板）。\n\n3.循环操作第2步。'

global myColor_win_bg
myColor_win_bg='white'

# --窗口初始化--
def windowInit():    
    str_title='程序标题：Excel另存为-zhy'

    myColor_transparent_color='#cccc99'    # 设个特殊颜色，作为透明颜色

    
    myColor_btn_fg='black'

    mainWindow = Tk()

# 主窗口设置
    mainWindow.title(str_title)
    mainWindow.attributes('-topmost',True)  # 窗口置顶

    mainWindow.overrideredirect(True)   # 不显示标题栏

    mainWindow.protocol('WM_DELETE_WINDOW',lambda:closeWindow(mainWindow))  # 绑定窗口关闭事件

    # config是设置控件属性；attributes是设置系统底层属性；mainWindow.config(bg=myColor_transparent_color)等效于mainWindow['bg']=myColor_transparent_color
    mainWindow.config(bg=myColor_transparent_color)
    mainWindow.attributes('-transparentcolor', myColor_transparent_color)  # 背景透明

    mainWindow.attributes('-alpha','0.9')   # 透明度

    width_win = 520
    height_win = 500

    x_win = (mainWindow.winfo_screenwidth() // 2) - (width_win // 2)
    y_win = (mainWindow.winfo_screenheight() // 2) - (height_win // 2)

    mainWindow.geometry('{}x{}+{}+{}'.format(width_win, height_win, x_win, y_win))  # 窗口居中，设置 窗口大小、位置：字符串格式：width x height + x + y

    mainWindow.resizable(False,False)   # 禁止修改窗口大小

# 工具栏框架
    fm_tool=Frame(master=mainWindow,bg=myColor_transparent_color)
    fm_tool.grid(row=1,column=1,columnspan=2,sticky='NW')

    # 标题画布
    width_canvas_title=width_win
    height_canvas_title=28

    myColor_tool_bg='gray'

    canvas_title=Canvas(master=fm_tool,width=width_canvas_title,height=height_canvas_title,highlightthickness=0,cursor='fleur')    # ,highlightthickness=0

    canvas_title.create_rectangle(0,0,width_canvas_title-1,height_canvas_title-1,outline=myColor_tool_bg,activefill=myColor_tool_bg)    # ,activefill=myColor_tool_bg
    
    canvas_title.create_text(1/2*width_canvas_title,1/2*height_canvas_title,text=str_title,fill=myColor_win_bg,state='disabled')    #,anchor='center'

    canvas_title.bind('<Button-1>',lambda event:funMouseDown(event))    # 绑定事件和函数
    canvas_title.bind('<B1-Motion>',lambda event:funMouseMove(event,mainWindow))

    canvas_title.grid(row=1,column=1,sticky='SW')

    # 另存为按钮画布
    width_canvas_btnsave=1/5*width_win
    height_canvas_btnsave=2*height_canvas_title

    # mainWindow.update_idletasks()   # 要先使用grid，再使用update_idletasks方法，最后用winfo_width方法，才能获取正确的width
    # width_canvas=canvas_btn_save.winfo_width()
    # height_canvas=canvas_btn_save.winfo_height()

    canvas_btn_save=Canvas(master=fm_tool,width=width_canvas_btnsave,height=height_canvas_btnsave,highlightthickness=0,cursor='hand2')    # ,highlightthickness=0
    
    canvas_btn_save.create_rectangle(0,0,width_canvas_btnsave-1,height_canvas_btnsave-1,fill=myColor_transparent_color,outline=myColor_transparent_color) #myColor_transparent_color

    myColor_btn_up='#1777FF'   #'#218868'
    myColor_btn_down='red'

    list_itemId=[]

    itemId=canvas_btn_save.create_rectangle(16,0,width_canvas_btnsave-17,height_canvas_btnsave-1,fill=myColor_btn_up,outline=myColor_btn_up) #myColor_transparent_color
    list_itemId.append(itemId)

    # 在点（x1，y1）和点（x2，y2）确定的矩形中画一个内接椭圆，以椭圆上与椭圆中心夹角为90°的点为起点，逆时针画一条角度为180°的默认style='pieslice'扇形，style:arc,chord,orpieslice
    itemId=canvas_btn_save.create_arc(0,0,1/3*width_canvas_btnsave+1,height_canvas_btnsave,start=90,extent=180,fill=myColor_btn_up,outline=myColor_btn_up)
    list_itemId.append(itemId)

    itemId=canvas_btn_save.create_arc(2/3*width_canvas_btnsave-1,0,width_canvas_btnsave-1,height_canvas_btnsave,start=90,extent=-180,fill=myColor_btn_up,outline=myColor_btn_up)

    list_itemId.append(itemId)

    canvas_btn_save.create_text(1/2*width_canvas_btnsave,1/2*height_canvas_btnsave,text='💾另存为',font=('SimHei',14),fill=myColor_win_bg,state='disabled')    # ,anchor='center','Microsoft Yahei'

    canvas_btn_save.grid(row=1,column=1,sticky='NW')

    # 关闭按钮
    btn_close=Button(master=fm_tool,text='X',font=('SimHei',12),relief='flat',bg='red',fg=myColor_win_bg)
    btn_close.grid(row=1,column=1,sticky='SE')
# 

# 左占位框架
    fm_left=Frame(master=mainWindow,bg=myColor_transparent_color)
    fm_left.grid(row=2,column=1,sticky='NW')
    Canvas(fm_left,width=1/2*width_canvas_btnsave,height=1,bg=myColor_transparent_color,highlightthickness=0).pack()
# 

# 右边的映射框架，包含 切换按钮、选择参数框架、配置框架
    fm_map=Frame(master=mainWindow)
    fm_map.grid(row=2,column=2,sticky='NW')
# 


# 切换框架按钮
    width_canvas_switch=width_win-1/2*width_canvas_btnsave
    height_canvas_switch=height_canvas_title

    canvas_switch=Canvas(master=fm_map,width=width_canvas_switch,height=height_canvas_switch,highlightthickness=0)

    canvas_switch.create_rectangle(0,0,width_canvas_switch-1,height_canvas_switch-1,outline=myColor_btn_fg,fill=myColor_tool_bg,activefill=myColor_btn_fg)
    canvas_switch.create_text(1/2*width_canvas_switch,1/2*height_canvas_switch,text='切换“选择参数”/“配置”',fill=myColor_win_bg,state='disabled') #,anchor='center'

    canvas_switch.bind('<Button-1>',lambda event:switchFrame(lbfm_select,lbfm_config))
    canvas_switch.grid(row=1,column=1,sticky='NW')
# 

# 选择参数框架
    global column_1
    column_1=1
    global column_2
    column_2=2

    lbfm_select=LabelFrame(master=fm_map,text='选择参数：',bg=myColor_win_bg)
    lbfm_select.grid(row=2,column=column_1,sticky='NW')    # ,columnspan=2

    myRow=1
    Label(lbfm_select,text='Excel源文件：',bg=myColor_win_bg).grid(row=myRow,column=1,sticky='NE')

    txt_srcFile=Text(lbfm_select,height=2,width=50,state='disabled',relief='solid')
    txt_srcFile.grid(row=myRow,column=2,sticky='NW',pady=3)

    btn_selectSrcFile=Button(lbfm_select,text='···',bg=myColor_win_bg,fg=myColor_btn_fg)
    btn_selectSrcFile.grid(row=myRow,column=3)

    myRow +=1
    lbl_targetFolder=Label(lbfm_select,text='保存到文件夹：',bg=myColor_win_bg)
    lbl_targetFolder.grid(row=myRow,column=1,sticky='NE')


    txt_targetFolder=Text(lbfm_select,height=2,width=50,relief='solid')
    
    txt_targetFolder.grid(row=myRow,column=2,sticky='NW',pady=3)

    Button(lbfm_select,text='···',bg=myColor_win_bg,fg=myColor_btn_fg,command=lambda : selectFileOrFolder(txt_targetFolder,isFile=False)).grid(row=myRow,column=3)

    myRow +=1
    Label(lbfm_select,text='运行中模板：',bg=myColor_win_bg).grid(row=myRow,column=1,sticky='NE')
    txt_TemplateFileName=Label(lbfm_select,text='运行中的-xxx.xlsx')
    txt_TemplateFileName.grid(row=myRow,column=2,sticky='WE')

    myRow +=1
    Label(lbfm_select,text='（自动生成“运行中模板”工作簿，在“保存到文件夹”下）').grid(row=myRow,column=2,sticky='WE')

    myRow +=1
    btn_showOrHideTips=Label(lbfm_select,text='隐藏 运行提示',relief='raised',borderwidth=1)
    btn_showOrHideTips.grid(row=myRow,column=1,sticky='SW')

# 

# 配置框架
    lbfm_config=LabelFrame(master=fm_map,text='配置：',bg=myColor_win_bg)
    lbfm_config.grid(row=2,column=column_2,sticky='NE')    # ,columnspan=2

    myRow=1
    lbl_sht_name=Label(lbfm_config,bg=myColor_win_bg)
    lbl_sht_name.grid(row=myRow,column=1,sticky='NE')

    txt_sht_name=Entry(lbfm_config,relief='solid')  # ,textvariable=StringVar(),width=30
    txt_sht_name.grid(row=myRow,column=2,sticky='NWE',padx=1,pady=1)

    myRow +=1
    lbl_filename_cellAddr=Label(lbfm_config,bg=myColor_win_bg)
    lbl_filename_cellAddr.grid(row=myRow,column=1,sticky='NE')

    txt_filename_cellAddr=Entry(lbfm_config,relief='solid') # width=30
    txt_filename_cellAddr.grid(row=myRow,column=2,sticky='NWE',padx=1,pady=1)

    myRow +=1
    lbl_orderNumber_cellAddr=Label(lbfm_config,bg=myColor_win_bg)
    lbl_orderNumber_cellAddr.grid(row=myRow,column=1,sticky='NE')

    txt_orderNumber_cellAddr=Entry(lbfm_config,relief='solid')
    txt_orderNumber_cellAddr.grid(row=myRow,column=2,sticky='NW',padx=1,pady=1)

    # myRow +=1
    # lbl_orderTime_cellAddr=Label(lbfm_config,text='制单时间单元格地址：',bg=myColor_win_bg)
    # lbl_orderTime_cellAddr.grid(row=myRow,column=1,sticky='NE')

    # txt_orderTime_cellAddr=Entry(lbfm_config,width=30,relief='solid')
    # txt_orderTime_cellAddr.grid(row=myRow,column=2,sticky='NW',padx=1,pady=1)

    myRow +=1
    btn_saveConfig=Button(lbfm_config,text='保存配置',bg=myColor_win_bg,fg=myColor_btn_fg)
    btn_saveConfig.grid(row=myRow,column=2,sticky='E')
# 

# 运行信息框架
    fm_runInfo=Frame(master=mainWindow)
    fm_runInfo.grid(row=3,column=2,sticky='NW')

    myRow =1

    txt_runInfo=Text(fm_runInfo,width=66,borderwidth=2,relief='ridge')
    txt_runInfo.insert('0.0',str_tips)

    txt_runInfo['height']=getStrWrapLines(str_tips,txt_runInfo['width'])

    txt_runInfo['state']='disabled'
    txt_runInfo.grid(row=myRow,column=1,sticky='NW')

    # txt_runInfo.pack(fill='y',expand='yes',side='left')

    # # 滚动条
    # scrollbar_v=Scrollbar(fm_runInfo,orient='vertical')
    # scrollbar_v.pack(fill='y',side='left')

    # txt_runInfo['yscrollcommand']=scrollbar_v.set
    # scrollbar_v['command']=txt_runInfo.yview
# 

# 读取配置，设定默认值
    list_mapping_value=readWriteMapping('r')    # 读取配置
    indexRow=0
    txt_targetFolder.insert('0.0',list_mapping_value[indexRow][1])
    txt_targetFolder['state']='disabled'
    indexRow+=1
    lbl_sht_name['text']=list_mapping_value[indexRow][0]
    txt_sht_name.insert('0',list_mapping_value[indexRow][1])
    indexRow+=1
    lbl_filename_cellAddr['text']=list_mapping_value[indexRow][0]
    txt_filename_cellAddr.insert('0',list_mapping_value[indexRow][1])
    indexRow+=1
    lbl_orderNumber_cellAddr['text']=list_mapping_value[indexRow][0]
    txt_orderNumber_cellAddr.insert('0',list_mapping_value[indexRow][1])


# 另存为按钮绑定事件
    dict_widget={'fm_map':fm_map,
                'fm_runInfo':fm_runInfo,
                'txt_runInfo':txt_runInfo,
                'txt_srcFile':txt_srcFile,
                'txt_targetFolder':txt_targetFolder,
                'txt_sht_name':txt_sht_name,
                'txt_filename_cellAddr':txt_filename_cellAddr,
                'txt_orderNumber_cellAddr':txt_orderNumber_cellAddr}
                # 'txt_orderTime_cellAddr':txt_orderTime_cellAddr


    dict_config1={'fill':myColor_btn_down,'outline':myColor_btn_down}   # 字典是引用类型,一个地方改变了值，其他地方读取的值就改变了
    canvas_btn_save.bind('<Button-1>',lambda event: myItemconfigure(canvas_btn_save,list_itemId,dict_config1))
    
    dict_config2={'fill':myColor_btn_up,'outline':myColor_btn_up}
    canvas_btn_save.bind('<ButtonRelease-1>',lambda event: [myItemconfigure(canvas_btn_save,list_itemId,dict_config2),workbookSaveas(dict_widget)])  # 绑定一个事件和多个函数 myItemconfigure,workbookSaveas

# 标题栏绑定显示映射框架的事件，防止后面运行程序出错框架不显示
    canvas_title.bind('<Double-Button-1>',lambda event:fm_map.grid(row=2,column=2,sticky='NW'))

# 源文件文本框绑定事件
    txt_srcFile.bind('<Button-3>',lambda event: openFolderOrFile(txt_srcFile))
# 保存文件夹文本框绑定事件
    txt_targetFolder.bind('<Button-3>',lambda event: openFolderOrFile(txt_targetFolder))
# 选择源文件按钮绑定事件
    btn_selectSrcFile.bind('<ButtonRelease-1>',lambda event: selectFileOrFolder(txt_srcFile,txt_TemplateFileName))

# 保存设置按钮绑定事件
    list_mapping=[[lbl_targetFolder,txt_targetFolder],
                [lbl_sht_name,txt_sht_name],
                [lbl_filename_cellAddr,txt_filename_cellAddr],
                [lbl_orderNumber_cellAddr,txt_orderNumber_cellAddr]]

    btn_saveConfig.bind('<ButtonRelease-1>',lambda event:[readWriteMapping('w',getConfigData(list_mapping)), messagebox.showinfo('配置已保存','配置已保存')])

# 关闭按钮绑定事件
    btn_close.bind('<ButtonRelease-1>',lambda event:[readWriteMapping('w',getConfigData(list_mapping)), closeWindow(mainWindow)])

# 显示隐藏运行信息
    btn_showOrHideTips.bind('<Button-1>',lambda event:showOrHideTips(btn_showOrHideTips,fm_runInfo))


# 显示窗口
    mainWindow.mainloop()


# # --函数：圆角矩形--
# def round_rectangle(myCanvas,x1, y1, x2, y2, radius=25, **kwargs):
#     # 左上点：x1,y1；右下点：x2,y2；radius：圆角距离两点的长度

#     points = [x1+radius, y1,  x1+radius, y1,     # 点重复两次平滑变成直线
#             x2-radius, y1,  x2-radius, y1,
#             x2, y1,
#             x2, y1+radius,  x2, y1+radius,
#             x2, y2-radius,  x2, y2-radius,
#             x2, y2,
#             x2-radius, y2,  x2-radius, y2,
#             x1+radius, y2,  x1+radius, y2,
#             x1, y2,
#             x1, y2-radius,  x1, y2-radius,
#             x1, y1+radius,  x1, y1+radius,
#             x1, y1]
#     # smooth=True 是关键：两点之间平滑
#     return myCanvas.create_polygon(points, **kwargs, smooth=True)



# --开始--
if __name__=='__main__':
    windowInit()
